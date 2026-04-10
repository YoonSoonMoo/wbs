import csv
import io

from app.extensions import get_db
from app.models import wbs_item as wbs_model

EXPORT_COLUMNS = [
    'wbs_code', 'category', 'task_name', 'subtask', 'detail',
    'assignee', 'plan_start', 'plan_end', 'actual_start', 'actual_end',
    'effort', 'progress', 'status',
]

EXPORT_HEADERS = [
    'WBS코드', '구분', 'Task', '서브태스크', '세부항목',
    '담당자', '계획시작', '계획완료', '실제시작', '실제종료',
    '공수', '진행률', '진행상태',
]


def export_csv(project_id):
    """WBS 항목을 CSV로 내보낸다. UTF-8 BOM 포함."""
    items = wbs_model.get_items_by_project(project_id)

    output = io.StringIO()
    output.write('\ufeff')  # BOM for Excel
    writer = csv.writer(output)
    writer.writerow(EXPORT_HEADERS)

    for item in items:
        row = [item.get(col, '') for col in EXPORT_COLUMNS]
        writer.writerow(row)

    return output.getvalue()


def export_excel(project_id):
    """WBS 항목을 Excel로 내보낸다."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    items = wbs_model.get_items_by_project(project_id)
    wb = Workbook()
    ws = wb.active
    ws.title = 'WBS'

    # 헤더 스타일
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')

    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    for row_idx, item in enumerate(items, 2):
        for col_idx, col in enumerate(EXPORT_COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(col, ''))

    # 컬럼 너비 자동 조정
    for col_idx, header in enumerate(EXPORT_HEADERS, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = max(len(header) * 2, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_csv(project_id, file_content):
    """CSV 파일 내용을 파싱하여 WBS 항목으로 가져온다."""
    reader = csv.DictReader(io.StringIO(file_content))
    header_map = dict(zip(EXPORT_HEADERS, EXPORT_COLUMNS))

    count = 0
    for row in reader:
        data = {'project_id': project_id}
        for header, value in row.items():
            col_name = header_map.get(header.strip())
            if col_name:
                data[col_name] = value.strip() if value else ''

        if data.get('task_name') or data.get('subtask') or data.get('detail'):
            data['sort_order'] = count
            wbs_model.create_item(data)
            count += 1

    return count


def import_paste_data(project_id, tsv_text):
    """탭 구분 데이터(Excel 붙여넣기)를 파싱하여 가져온다."""
    lines = tsv_text.strip().split('\n')
    if not lines:
        return 0

    count = 0
    for line in lines:
        cols = line.split('\t')
        if len(cols) < 3:
            continue

        data = {
            'project_id': project_id,
            'sort_order': count,
        }
        for idx, col_name in enumerate(EXPORT_COLUMNS):
            if idx < len(cols):
                data[col_name] = cols[idx].strip()

        if data.get('task_name') or data.get('subtask') or data.get('detail'):
            wbs_model.create_item(data)
            count += 1

    return count
